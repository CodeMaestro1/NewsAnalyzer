import sys
import os
import unittest
import json
from openpyxl import load_workbook
from mock import patch
from Project import ReadFiles, WriteFiles, JaccardIndex, BPlusTree


class TestReadFile(unittest.TestCase):

    def setUp(self):
        print("setUp was called before the unit test")


    def tearDown(self):
        print("tearDown was called after the unit test")

    def test_read_pairs_from_file(self):
        # Test case 1: Invalid file path
        file_to_read_categories_name_2 = r"\path\to\invalid\file.txt"
        self.bpt_instance = BPlusTree(10)
        expected_result = {}
        result = ReadFiles.read_pairs_from_file(file_to_read_categories_name_2, self.bpt_instance)
        self.assertEqual(result, expected_result)

        # Test case 2: Empty file
        file_to_read_categories_name_3 = r"C:\Users\user\Downloads\NewsAnalyzer-main\empty.txt"
        self.bpt_instance = BPlusTree(10)
        expected_result = {}
        result = ReadFiles.read_pairs_from_file(file_to_read_categories_name_3, self.bpt_instance)
        self.assertEqual(result, expected_result)

        # Test case 3: Valid file path
        file_to_read_categories_name = r"C:\Users\user\Downloads\NewsAnalyzer-main\category_docId.txt"
        self.bpt_instance = BPlusTree(10)

        expected_result = {
            "Category_A": {'1', '2', '3', '4'},
            "Category_B": {'5', '6', '7'},
            "Category_C": {'8', '9'}
        }

        result = ReadFiles.read_pairs_from_file(file_to_read_categories_name, self.bpt_instance)
        self.assertEqual(result, expected_result)



class TestJaccardIndex(unittest.TestCase):


    def setUp(self):
        print("setUp was called at the beginning of the unit test")
        #Dont forget to change this.
        #######################################################################################
        file_to_read_categories = r"C:\Users\user\Downloads\NewsAnalyzer-main\category_docId.txt"
        file_to_read_term = r"C:\Users\user\Downloads\NewsAnalyzer-main\docID_term.txt"
        file_to_read_stems = r"C:\Users\user\Downloads\NewsAnalyzer-main\stem_term.txt"
        #######################################################################################

        bpt_categories = BPlusTree(10)
        bpt__term = BPlusTree(10)
        bpt_stems = BPlusTree(10)
        returned_data_categories = ReadFiles.read_pairs_from_file(file_to_read_categories, bpt_categories)
        returned_data_term = ReadFiles.read_pairs_from_file(file_to_read_term, bpt__term)
        returned_data_stems = ReadFiles.read_pairs_from_file(file_to_read_stems, bpt_stems)
        self.jaccard_instance =JaccardIndex(returned_data_categories, returned_data_term, returned_data_stems)


    def tearDown(self):
        print("tearDown was called after the unit test")

    def test_calculate_jaccard_index(self):
        # Test case 1: The small set is a subset of the larger set
        result_dict = self.jaccard_instance.calculate_jaccard_index()
        small_dict = {
            'stem1': {'Category_A': 0.75, 'Category_B': 0, 'Category_C': 0},
            'stem30': {'Category_A': 0, 'Category_B': 0, 'Category_C': 0.5}
        }
        expected_result = all(item in result_dict.items() for item in small_dict.items())
        self.assertTrue(expected_result)

        # Test case 2: The small set is not a subset of the larger set
        small_dict = {
            'stem43': {'Category_A': 0.75, 'Category_B': 0, 'Category_C': 0},
            'stem30': {'Category_X': 0, 'Category_Y': 0, 'Category_Z': 0.5}
        }
        expected_result = all(item in result_dict.items() for item in small_dict.items())
        self.assertFalse(expected_result)


    def test_get_term_docs(self):
        # Test case 1: Empty term value
        term_value = []
        expected_result = set()
        result = self.jaccard_instance.get_term_docs(term_value)
        self.assertEqual(result, expected_result)

        # Test case 2: No matching documents
        result = {stem_key: set(self.jaccard_instance.get_term_docs(term_value)) for stem_key, term_value in self.jaccard_instance.stem.items()}
        small_dict = {
            'stem1': {'5', '6', '7'},
            'stem2': {'8', '9', '10'}
        }
        expected_result = all(item in result.items() for item in small_dict.items())
        self.assertFalse(expected_result)

        # Test case 3: Matching documents
        result = {stem_key: set(self.jaccard_instance.get_term_docs(term_value)) for stem_key, term_value in self.jaccard_instance.stem.items()}
        small_dict = {
            'stem1': {'1', '2', '3'},
            'stem2': {'1'},
            'stem33': {'9'}
        }
        expected_result = all(item in result.items() for item in small_dict.items())
        self.assertTrue(expected_result)

    def test_get_most_relevant_stems_for_category(self):
        # Test case 1: Existing category value
        self.jaccard_instance.calculate_jaccard_index()
        result = self.jaccard_instance.get_most_relevant_stems_for_category('Category_A', 3)
        expected_result = ['stem1', 'stem2', 'stem3']
        self.assertEqual(result, expected_result)

        # Test case 2: Non-existing category value
        result = self.jaccard_instance.get_most_relevant_stems_for_category('Category_X', 3)
        expected_result = []
        self.assertEqual(result, expected_result)

    def test_get_most_relevant_categories_for_stem(self):
        # Test case 1: Existing category value
        self.jaccard_instance.calculate_jaccard_index()
        result = self.jaccard_instance.get_most_relevant_categories_for_stem('stem1', 3)
        expected_result = ['Category_A', 'Category_B', 'Category_C']
        self.assertEqual(result, expected_result)

        # Test case 2: Non-existing category value
        result = self.jaccard_instance.get_most_relevant_categories_for_stem('stem3456', 3)
        expected_result = []
        self.assertEqual(result, expected_result)

class TestWriteFiles(unittest.TestCase):


    def setUp(self):
        print("setUp was called at the beginning of the unit test")
        #Dont forget to change this.
        #######################################################################################
        file_to_read_categories = r"C:\Users\user\Downloads\NewsAnalyzer-main\category_docId.txt"
        file_to_read_term = r"C:\Users\user\Downloads\NewsAnalyzer-main\docID_term.txt"
        file_to_read_stems = r"C:\Users\user\Downloads\NewsAnalyzer-main\stem_term.txt"
        #######################################################################################

        bpt_categories = BPlusTree(10)
        bpt__term = BPlusTree(10)
        bpt_stems = BPlusTree(10)
        returned_data_categories = ReadFiles.read_pairs_from_file(file_to_read_categories, bpt_categories)
        returned_data_term = ReadFiles.read_pairs_from_file(file_to_read_term, bpt__term)
        returned_data_stems = ReadFiles.read_pairs_from_file(file_to_read_stems, bpt_stems)
        self.jaccard_instance =JaccardIndex(returned_data_categories, returned_data_term, returned_data_stems)
        self.data = self.jaccard_instance.calculate_jaccard_index()


    def tearDown(self):
        print("tearDown was called after the unit test")

    @patch('builtins.print')
    def test_write_to_file(self, mock_print):
        # Test case 1: Valid file name but no file type
        write_files_instance = WriteFiles("ValidFile", self.data, "json")
        write_files_instance.type_of_file = None
        write_files_instance.write_to_file()
        mock_print.assert_called_with("File type or file name is not defined.")

        # Test case 2: Valid file type but no file name
        write_files_instance = WriteFiles("NoName", self.data, "json")
        write_files_instance.file_name = None
        write_files_instance.write_to_file()
        mock_print.assert_called_with("File type or file name is not defined.")

        # Test case 3: Valid file name but invalid file type
        write_files_instance = WriteFiles("ValidFile", self.data, "xlsx")
        write_files_instance.type_of_file = "dwg"
        write_files_instance.write_to_file()
        mock_print.assert_called_with(f"Unsupported file type: {write_files_instance.type_of_file}. Only json and xlsx are supported.")

    def test_write_to_json_file(self):
        # Test Case 1: Looking for existing data
        write_files_instance = WriteFiles("ValidFile1", self.data, "json")
        write_files_instance.write_to_file()
        self.assertTrue(os.path.exists("ValidFile1.json"))
        with open("ValidFile1.json", 'r') as f:
            actual_data = json.load(f)

        expected_data = [
            {"stem": "stem1", "category": "Category_A", "jaccardIndex": 0.75},
            {"stem": "stem21", "category": "Category_B", "jaccardIndex": 0.3333333333333333},
            {"stem": "stem36", "category": "Category_C", "jaccardIndex": 0.5}
        ]

        expected_result = all(item in actual_data for item in expected_data)
        self.assertTrue(expected_result)

        # Test Case 2: Looking for non-existing data
        expected_data = [
            {"stem": "stem113", "category": "Category_A", "jaccardIndex": 0.75},
            {"stem": "stem21", "category": "Category_Y", "jaccardIndex": 0.3333333333333333},
            {"stem": "stem360", "category": "Category_Z", "jaccardIndex": 0.5}
        ]
        expected_result = all(item in actual_data for item in expected_data)
        self.assertFalse(expected_result)

    def test_write_to_xlsx_file(self):
        # Test Case 1: Looking for existing data
        write_files_instance = WriteFiles("ValidFile2", self.data, "xlsx")
        write_files_instance.write_to_file()
        self.assertTrue(os.path.exists("ValidFile2.xlsx"))

        wb = load_workbook(filename = 'ValidFile2.xlsx')
        sheet = wb.active

        actual_data = []
        for row in sheet.iter_rows(values_only=True):
            actual_data.append(row)

        expected_data = [("stem1", "Category_A", 0.75), ("stem21", "Category_B", 0.3333333333333333), ("stem36", "Category_C", 0.5)]
        expected_result = all(item in actual_data for item in expected_data)
        self.assertTrue(expected_result)

        # Test Case 2: Looking for non-existing data
        expected_data = [("stem113", "Category_A", 0.75), ("stem21", "Category_Y", 0.3333333333333333), ("stem360", "Category_Z", 0.5)]
        expected_result = all(item in actual_data for item in expected_data)
        self.assertFalse(expected_result)


if __name__ == '__main__':
    sys.exit(unittest.main())
